# -*- coding: utf-8 -*-
"""
=======================================================
SCRAPER DATA RESEP MAKANAN COOKPAD INDONESIA
=======================================================
Install:
    pip install requests beautifulsoup4 pandas openpyxl lxml

Jalankan:
    python scraper_Cookpad.py
=======================================================
Mengambil data resep makanan dari https://cookpad.com/id
menggunakan requests + BeautifulSoup.
Data yang diambil:
  - Nama Resep
  - Bahan-bahan (Ingredients)
  - Langkah Memasak
  - Penulis / Author
  - Waktu Memasak
  - Jumlah Porsi
  - Deskripsi
  - URL Resep
  - Kategori Pencarian
=======================================================
"""

import os
import sys
import time
import random
import requests
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime
from urllib.parse import quote, urljoin

os.environ['PYTHONIOENCODING'] = 'utf-8'

# ===================== CONFIG =====================
BASE_URL = "https://cookpad.com"
SEARCH_URL = "https://cookpad.com/id/cari/{keyword}"
OUTPUT_CSV = "cookpad_resep_makanan.csv"
OUTPUT_EXCEL = "cookpad_resep_makanan.xlsx"
DELAY_MIN = 1.5        # Delay minimum (detik) antara request
DELAY_MAX = 3.0        # Delay maksimum (detik) antara request
MAX_RETRIES = 3        # Retry jika request gagal
MAX_PAGES_PER_KEYWORD = 3   # Jumlah halaman pencarian per keyword
MAX_RECIPES_TOTAL = 500     # Batas total resep yang diambil (sesuaikan)

# Daftar kategori makanan Indonesia yang akan di-scrape
KATEGORI_MAKANAN = [
    "nasi goreng", "rendang", "soto ayam", "gado gado", "bakso",
    "mie goreng", "ayam goreng", "sate ayam", "rawon", "opor ayam",
    "nasi uduk", "pecel lele", "gulai", "sambal", "sayur asem",
    "tempe goreng", "tahu goreng", "ikan bakar", "cap cay", "capcay",
    "tumis kangkung", "sop buntut", "pempek", "martabak", "klepon",
    "lontong sayur", "bubur ayam", "pepes ikan", "ayam bakar",
    "nasi kuning", "semur daging", "tongseng", "gudeg",
    "soto betawi", "es cendol", "kolak", "serabi",
    "pisang goreng", "perkedel", "lumpia", "risoles",
]
# ==================================================


def get_session() -> requests.Session:
    """Buat session dengan headers browser."""
    print("[*] Membuat session...")
    session = requests.Session()
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Referer": "https://cookpad.com/id",
    })

    # Akses homepage untuk set cookies
    try:
        resp = session.get(f"{BASE_URL}/id", timeout=30)
        resp.raise_for_status()
        print(f"    [OK] Session siap, cookies: {list(session.cookies.keys())}")
    except Exception as e:
        print(f"    [!] Warning saat akses homepage: {e}")

    return session


def delay_random():
    """Random delay agar tidak terdeteksi sebagai bot."""
    wait = random.uniform(DELAY_MIN, DELAY_MAX)
    time.sleep(wait)


def fetch_page(session: requests.Session, url: str) -> BeautifulSoup | None:
    """Fetch halaman dan return BeautifulSoup object."""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = session.get(url, timeout=30)
            if resp.status_code == 429:
                print(f"      [!] Rate limited (429), tunggu 10 detik...")
                time.sleep(10)
                continue
            resp.raise_for_status()
            return BeautifulSoup(resp.text, "lxml")

        except requests.exceptions.HTTPError as e:
            print(f"      [!] HTTP Error (attempt {attempt}): {e}")
            if attempt < MAX_RETRIES:
                time.sleep(DELAY_MAX * 2)
        except requests.exceptions.ConnectionError as e:
            print(f"      [!] Connection Error (attempt {attempt}): {e}")
            if attempt < MAX_RETRIES:
                time.sleep(DELAY_MAX * 3)
        except Exception as e:
            print(f"      [!] Error (attempt {attempt}): {e}")
            if attempt < MAX_RETRIES:
                time.sleep(DELAY_MAX * 2)
    return None


def get_recipe_links_from_search(session: requests.Session, keyword: str,
                                  max_pages: int = MAX_PAGES_PER_KEYWORD) -> list[dict]:
    """Ambil daftar link resep dari halaman pencarian Cookpad."""
    recipe_links = []
    encoded_keyword = quote(keyword)

    for page in range(1, max_pages + 1):
        if page == 1:
            url = SEARCH_URL.format(keyword=encoded_keyword)
        else:
            url = SEARCH_URL.format(keyword=encoded_keyword) + f"?page={page}"

        print(f"    Halaman {page}/{max_pages}: {url}")
        soup = fetch_page(session, url)
        if not soup:
            print(f"      [!] Gagal fetch halaman {page}")
            break

        # Cari semua link resep di halaman pencarian
        # Pattern: /id/resep/<id> 
        found_count = 0
        seen_urls = {r["url"] for r in recipe_links}

        for a_tag in soup.find_all("a", href=True):
            href = a_tag.get("href", "")
            # Filter link resep
            if "/id/resep/" in href and "/resep/baru" not in href:
                # Ambil judul dari teks link
                title = a_tag.get_text(strip=True)
                if not title or len(title) < 3:
                    continue
                # Skip link navigasi/utility
                if title.lower() in ("cetak", "edit resep", "hapus", "laporkan resep",
                                      "bagikan", "kirim cooksnap", "tulis resep",
                                      "lebih banyak", "tersimpan", "tambahkan ke folder",
                                      "lihat detail statistik"):
                    continue

                full_url = urljoin(BASE_URL, href)
                # Hapus query params
                full_url = full_url.split("?")[0]

                if full_url not in seen_urls:
                    recipe_links.append({
                        "url": full_url,
                        "title_preview": title,
                        "kategori": keyword,
                    })
                    seen_urls.add(full_url)
                    found_count += 1

        print(f"      [OK] Ditemukan {found_count} resep baru")

        if found_count == 0:
            # Tidak ada resep baru = mungkin sudah halaman terakhir
            break

        delay_random()

    return recipe_links


def parse_recipe_page(session: requests.Session, url: str, kategori: str) -> dict | None:
    """Parse halaman detail resep dan ambil semua informasi."""
    soup = fetch_page(session, url)
    if not soup:
        return None

    recipe = {
        "Nama Resep": "",
        "Deskripsi": "",
        "Bahan-bahan": "",
        "Jumlah Bahan": 0,
        "Langkah Memasak": "",
        "Jumlah Langkah": 0,
        "Waktu Memasak": "",
        "Jumlah Porsi": "",
        "Penulis": "",
        "Lokasi Penulis": "",
        "Kategori": kategori,
        "URL": url,
    }

    try:
        # === NAMA RESEP ===
        # Coba dari <title> tag
        title_tag = soup.find("title")
        if title_tag:
            title_text = title_tag.get_text(strip=True)
            # Format: "Resep XXX oleh YYY - Cookpad"
            if " oleh " in title_text:
                recipe["Nama Resep"] = title_text.split(" oleh ")[0].replace("Resep ", "", 1).strip()
            elif " - Cookpad" in title_text:
                recipe["Nama Resep"] = title_text.replace(" - Cookpad", "").replace("Resep ", "", 1).strip()

        # Fallback: cari h1 atau elemen dengan class recipe-title
        if not recipe["Nama Resep"]:
            h1 = soup.find("h1")
            if h1:
                recipe["Nama Resep"] = h1.get_text(strip=True)

        # === DESKRIPSI ===
        # Biasanya ada di og:description meta tag
        og_desc = soup.find("meta", attrs={"property": "og:description"})
        if og_desc and og_desc.get("content"):
            desc_text = og_desc["content"]
            # Hapus prefix "Resep XXX. "
            if ". " in desc_text:
                desc_text = desc_text.split(". ", 1)[1] if desc_text.startswith("Resep ") else desc_text
            recipe["Deskripsi"] = desc_text.strip()

        # === BAHAN-BAHAN ===
        ingredients = []
        # Cari section "Bahan-bahan" -- biasanya dalam <ol> atau <ul> setelah header
        ingredient_section = None
        for header in soup.find_all(["h2", "h3"]):
            if "Bahan" in header.get_text():
                ingredient_section = header
                break

        if ingredient_section:
            # Cari <ol> atau <ul> setelah header
            next_list = ingredient_section.find_next(["ol", "ul"])
            if next_list:
                for li in next_list.find_all("li"):
                    text = li.get_text(separator=" ", strip=True)
                    if text and len(text) > 1:
                        ingredients.append(text)

        # Fallback: cari dengan pattern ingredient di list items
        if not ingredients:
            for ol in soup.find_all("ol"):
                items = ol.find_all("li")
                # Ingredient list biasanya punya banyak item pendek
                if len(items) >= 3:
                    texts = [li.get_text(separator=" ", strip=True) for li in items if li.get_text(strip=True)]
                    # Cek apakah ini ingredient list (biasanya item pendek)
                    avg_len = sum(len(t) for t in texts) / max(len(texts), 1)
                    if avg_len < 80 and len(texts) >= 3:
                        ingredients = texts
                        break

        recipe["Bahan-bahan"] = " | ".join(ingredients) if ingredients else ""
        recipe["Jumlah Bahan"] = len(ingredients)

        # === LANGKAH MEMASAK ===
        steps = []
        step_section = None
        for header in soup.find_all(["h2", "h3"]):
            if "Cara Membuat" in header.get_text() or "Langkah" in header.get_text():
                step_section = header
                break

        if step_section:
            next_list = step_section.find_next(["ol", "ul"])
            if next_list:
                for li in next_list.find_all("li"):
                    text = li.get_text(separator=" ", strip=True)
                    # Filter out angka saja dan teks sangat pendek
                    if text and len(text) > 5 and not text.isdigit():
                        # Bersihkan duplikasi angka step di awal
                        import re as _re
                        cleaned = _re.sub(r'^\d+\s+', '', text).strip()
                        if cleaned and len(cleaned) > 5:
                            steps.append(cleaned)

        recipe["Langkah Memasak"] = " || ".join(steps) if steps else ""
        recipe["Jumlah Langkah"] = len(steps)

        # === PENULIS & LOKASI ===
        # Pattern di Cookpad: link ke /id/pengguna/<id> mengandung nama author
        for a_tag in soup.find_all("a", href=True):
            if "/id/pengguna/" in a_tag.get("href", ""):
                author_text = a_tag.get_text(separator=" ", strip=True)
                if author_text and len(author_text) > 1:
                    # Bisa ada username @xxx dan lokasi
                    parts = author_text.split("@")
                    recipe["Penulis"] = parts[0].strip()
                    if len(parts) > 1:
                        remaining = parts[1].strip()
                        # Cari lokasi setelah username
                        username_parts = remaining.split()
                        if len(username_parts) > 1:
                            recipe["Lokasi Penulis"] = " ".join(username_parts[1:]).strip()
                    break

        # === WAKTU MEMASAK & PORSI ===
        page_text = soup.get_text()
        # Cari pattern "XX menit" atau "XX jam"
        import re
        time_match = re.search(r'(\d+\s*(?:menit|jam|hari))', page_text, re.IGNORECASE)
        if time_match:
            recipe["Waktu Memasak"] = time_match.group(1).strip()

        # Cari pattern "X porsi"
        porsi_match = re.search(r'(\d+\s*porsi)', page_text, re.IGNORECASE)
        if porsi_match:
            recipe["Jumlah Porsi"] = porsi_match.group(1).strip()

    except Exception as e:
        print(f"      [!] Error parsing resep: {e}")
        return None

    # Validasi minimal: harus punya nama dan bahan
    if not recipe["Nama Resep"]:
        return None

    return recipe


def scrape_cookpad() -> list[dict]:
    """Fungsi utama: scrape resep dari Cookpad Indonesia."""
    print(f"\n{'='*70}")
    print(f"  SCRAPER DATA RESEP MAKANAN - COOKPAD INDONESIA")
    print(f"  Start: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*70}")
    print(f"  Kategori makanan: {len(KATEGORI_MAKANAN)} jenis")
    print(f"  Max halaman/keyword: {MAX_PAGES_PER_KEYWORD}")
    print(f"  Max total resep: {MAX_RECIPES_TOTAL}")
    print(f"{'='*70}\n")

    session = get_session()
    semua_resep = []
    seen_urls = set()

    print(f"\n[1] Mengumpulkan link resep dari pencarian...\n")

    all_recipe_links = []
    for idx, keyword in enumerate(KATEGORI_MAKANAN, 1):
        print(f"  [{idx}/{len(KATEGORI_MAKANAN)}] Mencari: '{keyword}'")
        links = get_recipe_links_from_search(session, keyword)
        new_links = [l for l in links if l["url"] not in seen_urls]
        for l in new_links:
            seen_urls.add(l["url"])
        all_recipe_links.extend(new_links)
        print(f"      Total link terkumpul: {len(all_recipe_links)}\n")

        if len(all_recipe_links) >= MAX_RECIPES_TOTAL * 2:
            print(f"  [INFO] Sudah cukup link ({len(all_recipe_links)}), lanjut ke scraping detail.\n")
            break

        delay_random()

    # Batasi jumlah link
    if len(all_recipe_links) > MAX_RECIPES_TOTAL:
        all_recipe_links = all_recipe_links[:MAX_RECIPES_TOTAL]

    print(f"\n{'='*70}")
    print(f"  [OK] Total {len(all_recipe_links)} link resep unik terkumpul")
    print(f"{'='*70}\n")

    print(f"[2] Mengambil detail setiap resep...\n")

    failed_count = 0
    for idx, link_info in enumerate(all_recipe_links, 1):
        url = link_info["url"]
        kategori = link_info["kategori"]

        pct = idx / len(all_recipe_links) * 100
        print(f"  [{idx}/{len(all_recipe_links)}] ({pct:.1f}%) {link_info['title_preview'][:50]}...",
              end=" ", flush=True)

        recipe = parse_recipe_page(session, url, kategori)

        if recipe:
            semua_resep.append(recipe)
            print(f"[OK] Bahan:{recipe['Jumlah Bahan']} Langkah:{recipe['Jumlah Langkah']}")
            failed_count = 0
        else:
            failed_count += 1
            print(f"[SKIP]")
            if failed_count >= 10:
                print(f"\n  [!] 10 kegagalan berturut-turut. Mungkin di-block. Tunggu 30 detik...")
                time.sleep(30)
                session = get_session()
                failed_count = 0

        # Backup setiap 50 resep
        if len(semua_resep) % 50 == 0 and len(semua_resep) > 0:
            print(f"\n  >>> Backup: {len(semua_resep)} resep...")
            try:
                df_backup = pd.DataFrame(semua_resep)
                df_backup.to_csv("cookpad_backup.csv", index=False, encoding="utf-8-sig")
                print(f"  >>> Backup tersimpan: cookpad_backup.csv\n")
            except Exception as e:
                print(f"  >>> Backup gagal: {e}\n")

        delay_random()

    print(f"\n{'='*70}")
    print(f"  [OK] Selesai scraping! Total: {len(semua_resep)} resep")
    print(f"  End: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*70}")

    return semua_resep


def simpan_csv(data: list[dict], filepath: str) -> None:
    """Simpan data ke CSV."""
    if not data:
        print("[!] Tidak ada data untuk disimpan")
        return

    df = pd.DataFrame(data)
    df.insert(0, "No", range(1, len(df) + 1))

    kolom_order = [
        "No", "Nama Resep", "Kategori", "Deskripsi", "Bahan-bahan",
        "Jumlah Bahan", "Langkah Memasak", "Jumlah Langkah",
        "Waktu Memasak", "Jumlah Porsi", "Penulis", "Lokasi Penulis", "URL",
    ]
    kolom_tersedia = [k for k in kolom_order if k in df.columns]
    df = df[kolom_tersedia]

    try:
        df.to_csv(filepath, index=False, encoding="utf-8-sig")
        print(f"    [OK] CSV tersimpan: {filepath}")
    except Exception as e:
        print(f"    [!] Error simpan CSV: {e}")


def simpan_excel(data: list[dict], filepath: str) -> None:
    """Simpan data ke Excel dengan format rapi."""
    if not data:
        print("[!] Tidak ada data untuk disimpan")
        return

    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("    [!] openpyxl belum terinstall. Simpan CSV saja.")
        return

    df = pd.DataFrame(data)
    df.insert(0, "No", range(1, len(df) + 1))

    kolom_order = [
        "No", "Nama Resep", "Kategori", "Deskripsi", "Bahan-bahan",
        "Jumlah Bahan", "Langkah Memasak", "Jumlah Langkah",
        "Waktu Memasak", "Jumlah Porsi", "Penulis", "Lokasi Penulis", "URL",
    ]
    kolom_tersedia = [k for k in kolom_order if k in df.columns]
    df = df[kolom_tersedia]

    print(f"\n[4] Menyimpan ke Excel: {filepath}...")

    try:
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Resep Cookpad")
            ws = writer.sheets["Resep Cookpad"]

            # ---- STYLING ----
            thin = Side(style="thin", color="CCCCCC")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            hdr_fill = PatternFill("solid", start_color="E85D26", end_color="E85D26")
            hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            alt_fill = PatternFill("solid", start_color="FFF3E0", end_color="FFF3E0")
            alt_fill_white = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
            c_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            l_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

            lebar_kolom = {
                "No": 6, "Nama Resep": 30, "Kategori": 18, "Deskripsi": 45,
                "Bahan-bahan": 55, "Jumlah Bahan": 12, "Langkah Memasak": 60,
                "Jumlah Langkah": 13, "Waktu Memasak": 14, "Jumlah Porsi": 12,
                "Penulis": 20, "Lokasi Penulis": 22, "URL": 40,
            }

            for col_idx, col_name in enumerate(kolom_tersedia, 1):
                width = lebar_kolom.get(col_name, 20)
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # Header styling (warna oranye khas Cookpad)
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = c_align
                cell.border = border
            ws.row_dimensions[1].height = 28

            # Data rows
            for row_idx in range(2, ws.max_row + 1):
                fill = alt_fill if row_idx % 2 == 0 else alt_fill_white
                ws.row_dimensions[row_idx].height = 22
                for col_idx in range(1, len(kolom_tersedia) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill
                    cell.border = border
                    cell.font = Font(name="Calibri", size=9)
                    col_name = kolom_tersedia[col_idx - 1]
                    if col_name in ("No", "Jumlah Bahan", "Jumlah Langkah",
                                    "Waktu Memasak", "Jumlah Porsi"):
                        cell.alignment = c_align
                    else:
                        cell.alignment = l_align

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        print(f"    [OK] Excel tersimpan: {filepath}")
    except Exception as e:
        print(f"    [!] Error simpan Excel: {e}")
        print(f"    [!] Fallback ke CSV...")
        simpan_csv(data, filepath.replace(".xlsx", "_fallback.csv"))


# ===================== MAIN =====================
if __name__ == "__main__":
    print("\n" + "=" * 70)
    print("  SCRAPER DATA RESEP MAKANAN COOKPAD INDONESIA")
    print("  Method: requests + BeautifulSoup (tanpa Selenium)")
    print("=" * 70)
    print(f"\n[INFO] Kategori makanan : {len(KATEGORI_MAKANAN)} jenis")
    print(f"[INFO] Max resep        : {MAX_RECIPES_TOTAL}")
    print(f"[INFO] Output           : {OUTPUT_CSV} & {OUTPUT_EXCEL}")
    print(f"[INFO] Delay            : {DELAY_MIN}-{DELAY_MAX} detik\n")

    # Mulai scraping
    data = scrape_cookpad()

    if data:
        # Simpan CSV
        print(f"\n[3] Menyimpan ke CSV: {OUTPUT_CSV}...")
        simpan_csv(data, OUTPUT_CSV)

        # Simpan Excel
        simpan_excel(data, OUTPUT_EXCEL)

        # Preview
        df = pd.DataFrame(data)
        print(f"\n{'='*70}")
        print("PREVIEW 5 RESEP PERTAMA:")
        print("=" * 70)
        for idx, row in df.head(5).iterrows():
            print(f"\n  [{idx+1}] {row['Nama Resep']}")
            print(f"      Kategori : {row['Kategori']}")
            print(f"      Bahan    : {row['Jumlah Bahan']} item")
            print(f"      Langkah  : {row['Jumlah Langkah']} step")
            print(f"      Waktu    : {row['Waktu Memasak']}")
            print(f"      Porsi    : {row['Jumlah Porsi']}")
            print(f"      Penulis  : {row['Penulis']}")

        # Statistik
        print(f"\n{'='*70}")
        print("STATISTIK:")
        print("=" * 70)
        print(f"  Total resep      : {len(data)}")
        print(f"  Rata-rata bahan  : {df['Jumlah Bahan'].mean():.1f} per resep")
        print(f"  Rata-rata langkah: {df['Jumlah Langkah'].mean():.1f} per resep")

        print(f"\n  Kategori (Top 10):")
        for kat, count in df['Kategori'].value_counts().head(10).items():
            print(f"    - {kat}: {count}")

        print(f"\n  Penulis terbanyak (Top 10):")
        for penulis, count in df['Penulis'].value_counts().head(10).items():
            print(f"    - {penulis}: {count}")

        print(f"\n{'='*70}")
        print(f"  [OK] SELESAI!")
        print(f"  File CSV   : {OUTPUT_CSV}")
        print(f"  File Excel : {OUTPUT_EXCEL}")
        print(f"  Total Data : {len(data)} resep makanan")
        print(f"{'='*70}")
    else:
        print("\n[!] TIDAK ADA DATA YANG BERHASIL DIAMBIL")
        print("[!] Periksa koneksi internet dan coba lagi")
