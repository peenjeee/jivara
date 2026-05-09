# -*- coding: utf-8 -*-
"""
=======================================================
SCRAPER SEMUA DATA OBAT BPOM - API Direct Method
=======================================================
Install:
    pip install requests pandas openpyxl

Jalankan:
    python scraper_bpom_all_data.py
=======================================================
Menggunakan API endpoint langsung dari cekbpom.pom.go.id
untuk mengambil seluruh data obat teregistrasi (~23.695 produk).
Jauh lebih cepat daripada Selenium karena tidak perlu browser.
=======================================================
"""

import os
import sys
import time
import json
import requests
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.environ['PYTHONIOENCODING'] = 'utf-8'

# ===================== CONFIG =====================
API_URL = "https://cekbpom.pom.go.id/produk-dt/01"
PAGE_SIZE = 1000          # Jumlah record per request (max ~1000)
OUTPUT_EXCEL = "obat_bpom_all_data.xlsx"
OUTPUT_CSV = "obat_bpom_all_data.csv"
BACKUP_EXCEL = "obat_bpom_backup.xlsx"
DELAY_BETWEEN_REQUESTS = 1.5  # Delay (detik) antara request, supaya tidak di-block
MAX_RETRIES = 3               # Retry jika request gagal
# ==================================================


def get_session() -> requests.Session:
    """Buat session dengan cookies dan headers yang sesuai."""
    print("[*] Membuat session dan mengambil cookies...")
    session = requests.Session()

    # Headers seperti browser biasa
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": "https://cekbpom.pom.go.id/produk-obat",
        "Origin": "https://cekbpom.pom.go.id",
    })

    # Pertama, akses halaman utama untuk mendapat cookies (XSRF-TOKEN, session)
    try:
        resp = session.get("https://cekbpom.pom.go.id/produk-obat", timeout=30)
        resp.raise_for_status()
        print(f"    [OK] Cookies diperoleh: {list(session.cookies.keys())}")

        # Ambil XSRF token dari cookies
        xsrf = session.cookies.get("XSRF-TOKEN", "")
        if xsrf:
            # Decode URL-encoded token
            from urllib.parse import unquote
            xsrf_decoded = unquote(xsrf)
            session.headers["X-XSRF-TOKEN"] = xsrf_decoded
            print(f"    [OK] XSRF-TOKEN ditemukan (panjang: {len(xsrf_decoded)})")

        # Cari CSRF meta tag dari HTML
        import re
        csrf_match = re.search(r'<meta\s+name="csrf-token"\s+content="([^"]+)"', resp.text)
        if csrf_match:
            csrf_token = csrf_match.group(1)
            session.headers["X-CSRF-TOKEN"] = csrf_token
            print(f"    [OK] CSRF-TOKEN ditemukan (panjang: {len(csrf_token)})")

    except Exception as e:
        print(f"    [!] Warning: Gagal mengambil cookies: {e}")
        print(f"    [!] Akan tetap coba request tanpa cookies...")

    return session


def build_form_data(draw: int, start: int, length: int, search: str = "") -> dict:
    """Bangun form data untuk DataTables server-side processing."""
    data = {
        "draw": str(draw),
        "start": str(start),
        "length": str(length),
        "search[value]": search,
        "search[regex]": "false",
        # Kolom-kolom DataTables
        "columns[0][data]": "DT_RowIndex",
        "columns[0][name]": "",
        "columns[0][searchable]": "false",
        "columns[0][orderable]": "false",
        "columns[0][search][value]": "",
        "columns[0][search][regex]": "false",
        "columns[1][data]": "PRODUCT_REGISTER",
        "columns[1][name]": "",
        "columns[1][searchable]": "true",
        "columns[1][orderable]": "false",
        "columns[1][search][value]": "",
        "columns[1][search][regex]": "false",
        "columns[2][data]": "PRODUCT_NAME",
        "columns[2][name]": "",
        "columns[2][searchable]": "true",
        "columns[2][orderable]": "false",
        "columns[2][search][value]": "",
        "columns[2][search][regex]": "false",
        "columns[3][data]": "REGISTRAR",
        "columns[3][name]": "",
        "columns[3][searchable]": "true",
        "columns[3][orderable]": "false",
        "columns[3][search][value]": "",
        "columns[3][search][regex]": "false",
        # Ordering
        "order[0][column]": "0",
        "order[0][dir]": "asc",
    }
    return data


def fetch_page(session: requests.Session, draw: int, start: int, length: int) -> dict:
    """Ambil satu halaman data dari API."""
    form_data = build_form_data(draw, start, length)

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = session.post(
                API_URL,
                data=form_data,
                timeout=60,
            )
            resp.raise_for_status()
            result = resp.json()

            # Validasi response
            if "data" not in result:
                print(f"      [!] Response tidak mengandung 'data' (attempt {attempt})")
                if attempt < MAX_RETRIES:
                    time.sleep(DELAY_BETWEEN_REQUESTS * 2)
                    continue
                return None

            return result

        except requests.exceptions.HTTPError as e:
            print(f"      [!] HTTP Error {resp.status_code} (attempt {attempt}): {e}")
            if resp.status_code == 419:  # CSRF token expired
                print("      [!] CSRF Token expired - mencoba refresh session...")
                session = get_session()
            if attempt < MAX_RETRIES:
                time.sleep(DELAY_BETWEEN_REQUESTS * 3)
                continue

        except requests.exceptions.ConnectionError as e:
            print(f"      [!] Connection Error (attempt {attempt}): {e}")
            if attempt < MAX_RETRIES:
                time.sleep(DELAY_BETWEEN_REQUESTS * 5)
                continue

        except json.JSONDecodeError as e:
            print(f"      [!] JSON Decode Error (attempt {attempt}): {e}")
            if attempt < MAX_RETRIES:
                time.sleep(DELAY_BETWEEN_REQUESTS * 2)
                continue

        except Exception as e:
            print(f"      [!] Error (attempt {attempt}): {e}")
            if attempt < MAX_RETRIES:
                time.sleep(DELAY_BETWEEN_REQUESTS * 2)
                continue

    return None


def parse_records(raw_data: list[dict]) -> list[dict]:
    """Parse raw API records ke format tabel yang diinginkan."""
    records = []

    for item in raw_data:
        try:
            record = {
                "No": item.get("DT_RowIndex", ""),
                "Nomor Registrasi": item.get("PRODUCT_REGISTER", "").strip(),
                "Nama Produk": item.get("PRODUCT_NAME", "").strip(),
                "Merk": item.get("PRODUCT_BRANDS", "-").strip(),
                "Komposisi": item.get("INGREDIENTS", "").strip(),
                "Bentuk Sediaan": item.get("PRODUCT_FORM", "").strip(),
                "Kemasan": item.get("PRODUCT_PACKAGE", "").strip(),
                "Pendaftar": item.get("REGISTRAR", "").strip(),
                "Tanggal Terbit": item.get("PRODUCT_DATE", "").strip(),
                "Masa Berlaku s/d": item.get("PRODUCT_EXPIRED", "").strip(),
                "Status": item.get("STATUS", "").strip(),
            }
            records.append(record)

        except Exception as e:
            continue

    return records


def scrape_all_drugs() -> list[dict]:
    """Scrape SEMUA data obat dari BPOM via API."""
    print(f"\n{'='*70}")
    print(f"  SCRAPER SEMUA DATA OBAT BPOM - API Direct Method")
    print(f"  Start: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*70}")

    session = get_session()
    semua_data = []
    draw = 1
    start = 0
    total_records = None

    # Pertama, ambil 1 record untuk tahu total
    print("\n[1] Mengecek total data obat...")
    test_result = fetch_page(session, draw=0, start=0, length=1)
    if test_result:
        total_records = test_result.get("recordsTotal", 0)
        print(f"    [OK] Total data obat: {total_records:,} produk")
    else:
        print("    [!] Gagal mendapat total - akan coba terus sampai habis")
        total_records = 999999

    total_pages = (total_records + PAGE_SIZE - 1) // PAGE_SIZE
    print(f"    [OK] Akan mengambil {total_pages} batch (masing-masing {PAGE_SIZE} record)\n")

    print("[2] Mulai mengambil data...\n")
    batch = 1
    failed_count = 0
    max_consecutive_fails = 5

    while start < total_records:
        remaining = total_records - start
        current_length = min(PAGE_SIZE, remaining)

        pct = (start / total_records * 100) if total_records > 0 else 0
        print(f"    Batch {batch}/{total_pages} "
              f"[{start+1:,}-{start+current_length:,} / {total_records:,}] "
              f"({pct:.1f}%)...", end=" ", flush=True)

        result = fetch_page(session, draw, start, current_length)

        if result is None:
            failed_count += 1
            print(f"[GAGAL] (fail #{failed_count})")
            if failed_count >= max_consecutive_fails:
                print(f"\n    [!] {max_consecutive_fails} kegagalan berturut-turut. Berhenti.")
                break
            # Skip ke batch berikutnya
            start += current_length
            draw += 1
            batch += 1
            time.sleep(DELAY_BETWEEN_REQUESTS * 3)
            continue

        failed_count = 0  # Reset counter
        records = parse_records(result.get("data", []))
        semua_data.extend(records)
        print(f"[OK] +{len(records)} (total: {len(semua_data):,})")

        # Backup setiap 5 batch
        if batch % 5 == 0:
            print(f"      >>> Backup: {len(semua_data):,} items...")
            try:
                df_backup = pd.DataFrame(semua_data)
                df_backup.to_csv(BACKUP_EXCEL.replace('.xlsx', '.csv'),
                                 index=False, encoding='utf-8-sig')
            except Exception as e:
                print(f"      [!] Backup gagal: {e}")

        # Update total jika berubah
        new_total = result.get("recordsTotal", total_records)
        if new_total != total_records:
            total_records = new_total
            total_pages = (total_records + PAGE_SIZE - 1) // PAGE_SIZE

        start += current_length
        draw += 1
        batch += 1

        # Delay sopan antara request
        time.sleep(DELAY_BETWEEN_REQUESTS)

    print(f"\n{'='*70}")
    print(f"  [OK] Selesai! Total: {len(semua_data):,} produk obat")
    print(f"  End: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*70}")

    return semua_data


def simpan_excel(data: list[dict], filepath: str) -> None:
    """Simpan data ke Excel dengan format tabel yang rapi."""
    if not data:
        print("[!] Tidak ada data untuk disimpan")
        return

    kolom_order = [
        "No",
        "Nomor Registrasi",
        "Nama Produk",
        "Merk",
        "Komposisi",
        "Bentuk Sediaan",
        "Kemasan",
        "Pendaftar",
        "Tanggal Terbit",
        "Masa Berlaku s/d",
        "Status",
    ]

    df = pd.DataFrame(data)

    # Re-number
    df["No"] = range(1, len(df) + 1)

    # Pastikan kolom sesuai urutan
    kolom_tersedia = [k for k in kolom_order if k in df.columns]
    df = df[kolom_tersedia]

    print(f"\n[3] Menyimpan ke Excel: {filepath}...")

    try:
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Obat BPOM")
            ws = writer.sheets["Obat BPOM"]

            # ---- STYLING ----
            thin = Side(style="thin", color="CCCCCC")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            hdr_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
            hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            alt_fill = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
            alt_fill_white = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
            c_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            l_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

            lebar_kolom = {
                "No": 6,
                "Nomor Registrasi": 22,
                "Nama Produk": 35,
                "Merk": 20,
                "Komposisi": 40,
                "Bentuk Sediaan": 25,
                "Kemasan": 40,
                "Pendaftar": 35,
                "Tanggal Terbit": 15,
                "Masa Berlaku s/d": 17,
                "Status": 12,
            }

            # Set lebar kolom
            for col_idx, col_name in enumerate(kolom_tersedia, 1):
                width = lebar_kolom.get(col_name, 20)
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # Header styling
            for col_idx, cell in enumerate(ws[1], 1):
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = c_align
                cell.border = border

            ws.row_dimensions[1].height = 28

            # Data rows styling
            for row_idx in range(2, ws.max_row + 1):
                fill = alt_fill if row_idx % 2 == 0 else alt_fill_white
                ws.row_dimensions[row_idx].height = 22

                for col_idx in range(1, len(kolom_tersedia) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill
                    cell.border = border
                    cell.font = Font(name="Calibri", size=9)

                    # Center alignment for No, Tanggal, Status
                    col_name = kolom_tersedia[col_idx - 1]
                    if col_name in ("No", "Tanggal Terbit", "Masa Berlaku s/d", "Status"):
                        cell.alignment = c_align
                    else:
                        cell.alignment = l_align

            # Freeze panes (header tetap terlihat saat scroll)
            ws.freeze_panes = "A2"

            # Auto-filter
            ws.auto_filter.ref = ws.dimensions

        print(f"    [OK] Excel tersimpan: {filepath}")
        print(f"    Total: {len(data):,} produk obat")

    except Exception as e:
        print(f"    [!] Error simpan Excel: {e}")
        # Fallback ke CSV
        print(f"    [!] Mencoba simpan sebagai CSV...")
        try:
            df.to_csv(filepath.replace('.xlsx', '.csv'), index=False, encoding='utf-8-sig')
            print(f"    [OK] CSV tersimpan: {filepath.replace('.xlsx', '.csv')}")
        except Exception as e2:
            print(f"    [!] CSV juga gagal: {e2}")


def simpan_csv(data: list[dict], filepath: str) -> None:
    """Simpan data ke CSV."""
    if not data:
        return

    df = pd.DataFrame(data)
    df["No"] = range(1, len(df) + 1)

    kolom_order = [
        "No", "Nomor Registrasi", "Nama Produk", "Merk", "Komposisi",
        "Bentuk Sediaan", "Kemasan", "Pendaftar", "Tanggal Terbit",
        "Masa Berlaku s/d", "Status",
    ]
    kolom_tersedia = [k for k in kolom_order if k in df.columns]
    df = df[kolom_tersedia]

    try:
        df.to_csv(filepath, index=False, encoding='utf-8-sig')
        print(f"    [OK] CSV tersimpan: {filepath}")
    except Exception as e:
        print(f"    [!] Error simpan CSV: {e}")


if __name__ == "__main__":
    print("\n" + "=" * 70)
    print("  SCRAPER SEMUA DATA OBAT BPOM")
    print("  Method: API Direct (tanpa Selenium/Browser)")
    print("=" * 70)
    print(f"\n[INFO] Estimasi waktu: ~1-3 menit untuk ~23.695 produk")
    print(f"[INFO] Output: {OUTPUT_EXCEL} dan {OUTPUT_CSV}")
    print(f"[INFO] Delay antar request: {DELAY_BETWEEN_REQUESTS} detik\n")

    # Mulai scraping
    data = scrape_all_drugs()

    if data:
        # Simpan Excel
        simpan_excel(data, OUTPUT_EXCEL)

        # Simpan CSV juga
        simpan_csv(data, OUTPUT_CSV)

        # Preview
        df = pd.DataFrame(data)
        print(f"\n{'='*70}")
        print("PREVIEW 10 DATA PERTAMA:")
        print("=" * 70)
        preview_cols = ["Nomor Registrasi", "Nama Produk", "Komposisi", "Bentuk Sediaan", "Pendaftar"]
        preview_df = df[preview_cols].head(10)
        for idx, row in preview_df.iterrows():
            print(f"\n  [{idx+1}] {row['Nama Produk']}")
            print(f"      Reg: {row['Nomor Registrasi']}")
            print(f"      Komposisi: {row['Komposisi']}")
            print(f"      Bentuk: {row['Bentuk Sediaan']}")
            print(f"      Pendaftar: {row['Pendaftar']}")

        # Statistik
        print(f"\n{'='*70}")
        print("STATISTIK:")
        print("=" * 70)
        print(f"  Total produk: {len(data):,}")
        print(f"\n  Bentuk Sediaan (Top 10):")
        for bentuk, count in df['Bentuk Sediaan'].value_counts().head(10).items():
            print(f"    - {bentuk}: {count:,}")

        print(f"\n  Status:")
        for status, count in df['Status'].value_counts().items():
            print(f"    - {status}: {count:,}")

        print(f"\n  Pendaftar terbanyak (Top 10):")
        for pendaftar, count in df['Pendaftar'].value_counts().head(10).items():
            print(f"    - {pendaftar}: {count:,}")

        print(f"\n{'='*70}")
        print(f"  [OK] SELESAI!")
        print(f"  File Excel : {OUTPUT_EXCEL}")
        print(f"  File CSV   : {OUTPUT_CSV}")
        print(f"  Total Data : {len(data):,} produk obat")
        print(f"{'='*70}")
    else:
        print("\n[!] TIDAK ADA DATA YANG BERHASIL DIAMBIL")
        print("[!] Periksa koneksi internet dan coba lagi")
